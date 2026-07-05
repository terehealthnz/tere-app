#!/usr/bin/env python3
"""
build_pharmacy_list.py

Downloads the Medsafe Pharmacy Licence Register — the official monthly list of
every licensed pharmacy in New Zealand — and converts it into pharmacies.json
(served from public/, so it lands at tere.co.nz/pharmacies.json) and a
pharmacies.csv artefact in the repo root.

Source page: https://www.medsafe.govt.nz/compliance/pharmacy-licence-register.asp
The register is published monthly and contains, per pharmacy:
  Legal Entity (operator), Premises (trading) Name, Street address, Region.
It does NOT include phone/email or opening hours.

Usage:
  python build_pharmacy_list.py                       # download + convert
  python build_pharmacy_list.py --source file.xlsx    # convert a file already downloaded
  python build_pharmacy_list.py --json-out public/pharmacies.json --csv-out pharmacies.csv

Safety guards (run unattended from cron / GitHub Actions):
  · Aborts with non-zero exit + clear message if new list has fewer than 500 records.
  · Aborts if new list is more than 20% smaller than the existing pharmacies.json
    already committed. Prevents a broken download or a Medsafe format change from
    silently overwriting a good list.

Dependencies:
  pip install requests openpyxl
"""

import argparse
import csv
import io
import json
import re
import sys
from pathlib import Path

REGISTER_URL = "https://www.medsafe.govt.nz/compliance/PharmacyLicenceRegister.xlsx"

# Guardrails (see docstring above).
MIN_RECORDS = 500
MAX_SHRINK_FRACTION = 0.20  # abort if new list is >20% smaller than existing

# Header keywords -> normalised field name. Matched case-insensitively against
# the column heading with underscores swapped for spaces first, so the current
# Medsafe schema (LEGAL_ENTITY_NAME / PREMISES_NAME / STREET_ADDRESS /
# HEALTHNZ_DISTRICT / HEALTHNZ_REGION) matches as well as older whitespace-titled
# releases. Order matters: earlier hints win, so `district` (specific area) is
# picked for the town column before `region` (broader HNZ zone).
FIELD_HINTS = [
    ("legal_entity", ["legal entity", "operator", "licensee"]),
    ("premises_name", ["premises name", "trading", "pharmacy name"]),
    ("street", ["street address", "street", "address"]),
    ("suburb", ["suburb"]),
    ("town", ["healthnz district", "district", "town", "city"]),
    ("postcode", ["post code", "postcode", "postal"]),
    ("region", ["healthnz region", "region", "dhb"]),
]


def download(url: str) -> bytes:
    import requests
    headers = {"User-Agent": "Mozilla/5.0 (pharmacy-list-builder)"}
    resp = requests.get(url, headers=headers, timeout=60)
    resp.raise_for_status()
    return resp.content


def load_rows(xlsx_bytes: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [[("" if c is None else str(c).strip()) for c in row]
            for row in ws.iter_rows(values_only=True)]


def find_header(rows):
    """Return (header_row_index, {normalised_field: column_index}). Underscores
    in column titles are treated as spaces so LEGAL_ENTITY_NAME matches
    "legal entity"."""
    for i, row in enumerate(rows[:15]):
        norm = [c.lower().replace("_", " ") for c in row]
        mapping = {}
        for field, hints in FIELD_HINTS:
            for col, cell in enumerate(norm):
                if cell and field not in mapping and any(h in cell for h in hints):
                    mapping[field] = col
        # Treat as the header row once the anchor columns are found.
        if "premises_name" in mapping or ("legal_entity" in mapping and "street" in mapping):
            return i, mapping
    raise SystemExit(
        "Could not locate the header row. Open the .xlsx, check the column titles, "
        "then adjust FIELD_HINTS at the top of this script."
    )


def slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def build_records(rows, header_idx, cols):
    records, seen = [], set()
    for row in rows[header_idx + 1:]:
        def get(field):
            idx = cols.get(field)
            return row[idx].strip() if idx is not None and idx < len(row) else ""

        premises = get("premises_name") or get("legal_entity")
        if not premises:
            continue  # skip blank / spacer rows

        address = ", ".join(p for p in (get("street"), get("suburb"),
                                        get("town"), get("postcode")) if p)
        town = get("town")

        key = (premises.lower(), address.lower())
        if key in seen:
            continue
        seen.add(key)

        rec_id = slugify(f"{premises}-{town}") or slugify(premises) or str(len(records))
        records.append({
            "id": rec_id,
            "premises_name": premises,
            "legal_entity": get("legal_entity"),
            "address": address,
            "town": town,
            "region": get("region"),
        })
    records.sort(key=lambda r: (r["region"].lower(), r["premises_name"].lower()))
    return records


def write_json(records, json_path: Path):
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")


def write_csv(records, csv_path: Path):
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["id", "premises_name", "legal_entity", "address", "town", "region"]
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(records)


def validate_or_abort(records, existing_json_path: Path) -> None:
    """Guardrails against a broken download or a Medsafe format change silently
    overwriting a good list. Exits non-zero with a clear message on failure."""
    new_count = len(records)
    if new_count < MIN_RECORDS:
        sys.exit(
            f"ABORT: new list has only {new_count} records — below the {MIN_RECORDS} "
            f"minimum. Refusing to overwrite the existing file. Check that the .xlsx "
            f"downloaded correctly and that Medsafe hasn't changed the column layout."
        )
    if existing_json_path.exists():
        try:
            existing = json.loads(existing_json_path.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"WARNING: could not read existing {existing_json_path} for comparison ({e}). "
                  f"Skipping shrink check.")
            return
        existing_count = len(existing) if isinstance(existing, list) else 0
        if existing_count > 0:
            shrink = (existing_count - new_count) / existing_count
            if shrink > MAX_SHRINK_FRACTION:
                pct = int(shrink * 100)
                sys.exit(
                    f"ABORT: new list ({new_count}) is {pct}% smaller than the existing "
                    f"list ({existing_count}). Threshold is {int(MAX_SHRINK_FRACTION * 100)}%. "
                    f"Refusing to overwrite — check the source for a truncated download or "
                    f"a Medsafe schema change before re-running."
                )


def flag_hospital_pharmacies(records) -> None:
    """Medsafe doesn't cleanly flag hospital vs community pharmacies. Surface any
    entries whose legal entity or premises name suggests a hospital / DHB / Health NZ
    context so the human operator can decide what to do — this script never filters
    them silently."""
    keywords = [
        "hospital", "dhb", "district health board", "health new zealand",
        "te whatu ora", "district health services",
    ]
    def is_hospitalish(rec) -> bool:
        blob = " ".join([
            rec.get("premises_name", "") or "",
            rec.get("legal_entity", "") or "",
            rec.get("address", "") or "",
        ]).lower()
        return any(k in blob for k in keywords)
    flagged = [r for r in records if is_hospitalish(r)]
    if not flagged:
        return
    print(f"\n⚠ {len(flagged)} entries look hospital-affiliated (review manually — not filtered):")
    for r in flagged[:40]:
        print(f"   · {r['premises_name']}  ({r.get('legal_entity','')})  [{r.get('region','')}]  {r['id']}")
    if len(flagged) > 40:
        print(f"   … and {len(flagged) - 40} more.")


def main():
    ap = argparse.ArgumentParser(
        description="Build an NZ pharmacy pick-list from the Medsafe licence register.")
    ap.add_argument("--source", help="Path to a PharmacyLicenceRegister.xlsx already downloaded.")
    ap.add_argument("--json-out", default="public/pharmacies.json",
                    help="JSON output path (default: public/pharmacies.json — served as /pharmacies.json).")
    ap.add_argument("--csv-out", default="pharmacies.csv",
                    help="CSV artefact path (default: pharmacies.csv in repo root).")
    args = ap.parse_args()

    if args.source:
        xlsx_bytes = Path(args.source).read_bytes()
    else:
        print(f"Downloading {REGISTER_URL} ...")
        try:
            xlsx_bytes = download(REGISTER_URL)
        except Exception as e:
            sys.exit(f"Download failed ({e}).\n"
                     f"Download the file manually from the register page, then re-run:\n"
                     f"  python build_pharmacy_list.py --source PharmacyLicenceRegister.xlsx")

    rows = load_rows(xlsx_bytes)
    header_idx, cols = find_header(rows)
    records = build_records(rows, header_idx, cols)

    json_path = Path(args.json_out)
    csv_path = Path(args.csv_out)

    # Guardrails BEFORE writing files — nothing gets overwritten if this exits.
    validate_or_abort(records, json_path)

    write_json(records, json_path)
    write_csv(records, csv_path)
    print(f"Wrote {len(records)} pharmacies to {json_path} and {csv_path}")

    flag_hospital_pharmacies(records)


if __name__ == "__main__":
    main()
